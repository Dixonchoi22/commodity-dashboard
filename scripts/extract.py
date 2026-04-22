"""Extract report source files into JSON for a given period.

Reads raw sources from data/{period}/raw/:
  - overview.pdf      Expana monthly commodity price change report
  - hicp.xlsx         Eurostat HICP food inflation by country
  - forecast.xlsx     Mintec forward curves

Writes JSON under data/{period}/:
  - commodities.json  Summary table (category, name, MoM, YoY) for all commodities
  - commentary.json   Per-commodity price line + narrative paragraphs
  - hicp.json         EU and country-level HICP food inflation YoY by month
  - forecast.json     Daily forward curves

Usage:
  python scripts/extract.py 2026-04
"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent


def paths_for(period: str) -> dict[str, Path]:
    base = ROOT / "data" / period
    return {
        "base": base,
        "raw": base / "raw",
        "pdf": base / "raw" / "overview.pdf",
        "hicp_xlsx": base / "raw" / "hicp.xlsx",
        "forecast_xlsx": base / "raw" / "forecast.xlsx",
    }

CATEGORIES = [
    "Grains & Feed",
    "Oilseeds & Vegetable Oils",
    "Meat & Poultry",
    "Dairy & Eggs",
    "Fish & Seafood",
    "Fruit & Vegetables",
    "Juices",
    "Softs",
    "Herbs & Spices",
    "Nuts & Dried Fruit",
    "Textiles",
    "Packaging",
]


def pdf_to_text(pdf_path: Path) -> str:
    """Full-page text (columns will interleave) — used for the summary
    table on pages 1-4, which is easier to parse with both columns visible."""
    return subprocess.check_output(
        ["pdftotext", "-layout", str(pdf_path), "-"], text=True, encoding="utf-8"
    )


def pdf_columns_to_text(pdf_path: Path) -> str:
    """Return left-column + right-column text concatenated. We crop the PDF
    to each half separately so pdftotext doesn't interleave the two columns'
    paragraphs — critical for the per-commodity commentary sections."""
    # A4 page: 841.92 x 595.2 pts. The two-column layout puts each column
    # in ~420pt wide bands. Small overlap (5pt) absorbs borderline text.
    left = subprocess.check_output(
        ["pdftotext", "-layout", "-x", "0", "-y", "0", "-W", "425", "-H", "600", str(pdf_path), "-"],
        text=True, encoding="utf-8",
    )
    right = subprocess.check_output(
        ["pdftotext", "-layout", "-x", "416", "-y", "0", "-W", "425", "-H", "600", str(pdf_path), "-"],
        text=True, encoding="utf-8",
    )
    return left + "\n\n----COLUMN-BREAK----\n\n" + right


# ---------- Summary table extraction ----------

PCT = re.compile(r"([+\-]?\d+(?:\.\d+)?)%")
ROW_2PCT = re.compile(r"([+\-]?\d+(?:\.\d+)?%)\s+([+\-]?\d+(?:\.\d+)?%)")


def parse_pct(s: str) -> float | None:
    m = PCT.search(s)
    return float(m.group(1)) if m else None


def _load_canonical_categories() -> dict[str, str]:
    """Canonical name -> category map, extracted from the Sep 2025 legacy
    dashboard (which Gemini categorised manually). Used to fix the flaky
    category detection from the two-column PDF layout."""
    path = ROOT / "data" / "_canonical_categories.json"
    if path.exists():
        return json.loads(path.read_text())
    return {}


def extract_summary(text: str) -> list[dict]:
    """Parse the 4-page summary table into {category, name, mom, yoy} rows.

    Layout: two side-by-side columns. Each row ends with `...MoM%  YoY%`.
    Category labels appear centered in a separate mini-column. We detect
    category via a canonical name lookup first (reliable) and only fall
    back to positional detection when the name is new.
    """
    canonical = _load_canonical_categories()
    lines = text.splitlines()
    rows: list[dict] = []

    # Category runs: iterate through lines; when a category appears at the start
    # of a column, remember it for that column.
    # Simpler approach: scan for rows with two percentages. Split line at col
    # midpoint to pick name. Track category via preceding category-labeled rows.
    left_cat: str | None = None
    right_cat: str | None = None

    # We'll find the column midpoint by locating the gap between columns on
    # the header line(s). Hardcode to ~80 based on layout.
    MID = 80

    for line in lines:
        stripped = line.rstrip()
        if not stripped:
            continue
        # Category detection: look for exact category tokens anywhere on line
        for cat in CATEGORIES:
            if cat in line:
                col = line.index(cat)
                if col < MID:
                    left_cat = cat
                else:
                    right_cat = cat
        # Iterate the MoM/YoY pairs on this line. For each pair, the commodity
        # name is the text BETWEEN the previous pair's end and this pair's
        # start (or line start if it's the first pair). This avoids letting a
        # right-column row absorb the left-column row's name + percentages.
        prev_end = 0
        for m in ROW_2PCT.finditer(line):
            segment = line[prev_end:m.start()]
            prev_end = m.end()

            name = segment
            # Strip any full category label that appears within the segment
            for c in CATEGORIES:
                if c in name:
                    name = name.replace(c, " ")
            # Some multi-word category labels wrap across pdftotext lines,
            # leaving fragments like "Oilseeds & Vegetable" or "Oils" attached
            # to the first / last row of the section. Strip those too.
            for frag in (
                "Oilseeds & Vegetable",
                "Oilseeds &amp; Vegetable",
                "Fruit & Vegetables",
                "Dairy & Eggs",
                "Meat & Poultry",
                "Fish & Seafood",
                "Nuts & Dried Fruit",
                "Herbs & Spices",
                "Grains & Feed",
            ):
                name = name.replace(frag, " ")
            # Leading stray words left behind by wrapped categories
            name = name.strip()
            name = re.sub(r"^(Oils|Vegetable|Vegetables|Spices|Fruit|Eggs|Poultry|Seafood|Feed|Dried)\s+", "", name)
            # Strip PDF page-footer / header noise that pdftotext occasionally
            # pastes next to a commodity name ("2 Expana © 2026", etc.)
            name = re.sub(r"\b\d+\s*Expana\s*©\s*\d{4}\b", "", name)
            name = re.sub(r"\bExpana\s*©\s*\d{4}\b", "", name)
            name = re.sub(r"\bContents\b", "", name)
            name = " ".join(name.split())

            # Decide which column we're on by the match's x position
            col = m.start()
            cat = left_cat if col < MID else right_cat

            # Skip noise: empty name, header row, stray % markers
            if not name or name.lower().startswith("commodity"):
                continue
            if any(tok in name for tok in ("%", "CATEGORY", "Price changes")):
                continue

            # Prefer canonical category (name-based lookup) over positional
            # detection. Falls back to positional cat if name is unseen.
            canonical_cat = canonical.get(name.lower())
            if not canonical_cat:
                # Try a loose match (first 2-3 tokens) to catch rewordings like
                # "NFDM - SMP US" vs "SMP US".
                key = name.lower()
                for canon_name, canon_cat in canonical.items():
                    if canon_name in key or key in canon_name:
                        canonical_cat = canon_cat
                        break
            mom_s, yoy_s = m.group(1), m.group(2)
            rows.append(
                {
                    "category": canonical_cat or cat or "Unknown",
                    "name": name,
                    "mom_pct": parse_pct(mom_s),
                    "yoy_pct": parse_pct(yoy_s),
                }
            )
    # Dedupe (same row might appear if layout detection double-matches)
    seen = set()
    out: list[dict] = []
    for r in rows:
        key = (r["category"], r["name"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


# ---------- Per-commodity commentary extraction ----------

PRICE_LINE_PATTERNS = [
    # Standard: "The average weekly price of X as of March 25 increased by
    # 2.2%\s*MOM and declined by 0.2%\s*YOY to €500.05/MT"
    re.compile(
        r"The average weekly(?:\s+spot)? price of (?P<name>[A-Z][^.]+?) as of (?P<as_of>[A-Za-z]+ \d+) "
        r"(?P<dir1>increased|decreased|rose|declined|fell) by (?P<mom>\d+(?:\.\d+)?)%\s*MOM "
        r"and (?P<dir2>increased|decreased|rose|declined|fell) by (?P<yoy>\d+(?:\.\d+)?)%\s*YOY "
        r"to (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)",
        re.IGNORECASE,
    ),
    # Unchanged MOM: "The average weekly price of UK Butter as of March 25
    # was unchanged MOM and declined by 40.0%\s*YOY to £3,750/MT"
    re.compile(
        r"The average weekly(?:\s+spot)? price of (?P<name>[A-Z][^.]+?) as of (?P<as_of>[A-Za-z]+ \d+) "
        r"was unchanged MOM "
        r"and (?P<dir2>increased|decreased|rose|declined|fell) by (?P<yoy>\d+(?:\.\d+)?)%\s*YOY "
        r"to (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)",
        re.IGNORECASE,
    ),
    # "In February, the average price of Norwegian Haddock increased by..."
    # Also matches looser modifiers like "the average 3m futures price of
    # LME Aluminium..." via the optional modifier-words group.
    re.compile(
        r"In (?P<as_of>[A-Za-z]+), the average(?:\s+[a-zA-Z0-9 \-()]+?)? price of (?P<name>[A-Z][^.]+?) "
        r"(?P<dir1>increased|decreased|rose|declined|fell) by (?P<mom>\d+(?:\.\d+)?)%\s*MOM "
        r"and (?P<dir2>increased|decreased|rose|declined|fell) by (?P<yoy>\d+(?:\.\d+)?)%\s*YOY "
        r"to (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)",
        re.IGNORECASE,
    ),
    # Chained second clause: "..., the average price of EU Cartonboard GT2
    # was unchanged MOM and declined by 7.1% YOY to €785/MT"
    re.compile(
        r",\s*the average(?:\s+[a-zA-Z0-9 \-()]+?)? price of (?P<name>[A-Z][^.]+?) "
        r"was unchanged MOM "
        r"and (?P<dir2>increased|decreased|rose|declined|fell) by (?P<yoy>\d+(?:\.\d+)?)%\s*YOY "
        r"to (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)",
        re.IGNORECASE,
    ),
    # Unchanged MOM + "In Month" variant
    re.compile(
        r"In (?P<as_of>[A-Za-z]+), the average(?:\s+weekly)? price of (?P<name>[A-Z][^.]+?) "
        r"was unchanged MOM "
        r"and (?P<dir2>increased|decreased|rose|declined|fell) by (?P<yoy>\d+(?:\.\d+)?)%\s*YOY "
        r"to (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)",
        re.IGNORECASE,
    ),
    # Expana Benchmark (vanilla): "The Expana Benchmark Prices (EBP) of
    # vanilla from Madagascar ... were unchanged MOM in March, down 37.8%
    # YOY, at $28/kg..."
    re.compile(
        r"The Expana Benchmark Prices \(EBP\) of (?P<name>[^,.]+?),[^.]*?"
        r"were unchanged MOM in (?P<as_of>[A-Za-z]+),?\s*"
        r"(?P<dir2>down|up) (?P<yoy>\d+(?:\.\d+)?)%\s*YOY,?\s*"
        r"at (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)",
        re.IGNORECASE,
    ),
    # Nuts EBP: "The {Origin} {nut} EBP increased by 3.2% MOM in March to
    # $3.21/lb, down 1.2% YOY." (cashew, hazelnut, brazil, apricot, sultana)
    re.compile(
        r"The (?P<name>[A-Z][^.]+? EBP(?: \([^)]+\))?) "
        r"(?P<dir1>increased|decreased|rose|declined|fell) by (?P<mom>\d+(?:\.\d+)?)%\s*MOM "
        r"in (?P<as_of>[A-Za-z]+) "
        r"to (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)"
        r"(?:[,.]\s*(?:and\s+)?(?P<dir2>rose|declined|fell|increased|decreased|up|down)\s*(?:by\s*)?"
        r"(?P<yoy>\d+(?:\.\d+)?)%\s*YOY)?",
        re.IGNORECASE,
    ),
    # EBP with trailing price: "The Turkish sultana EBP (CIF NW Europe)
    # fell by 3.3% MOM in March, down 16% YOY, to $2,647/MT"
    re.compile(
        r"The (?P<name>[A-Z][^.]+? EBP(?: \([^)]+\))?) "
        r"(?P<dir1>increased|decreased|rose|declined|fell) by (?P<mom>\d+(?:\.\d+)?)%\s*MOM "
        r"in (?P<as_of>[A-Za-z]+),?\s*"
        r"(?P<dir2>up|down|rose|declined|fell|increased|decreased) (?:by\s*)?(?P<yoy>\d+(?:\.\d+)?)%\s*YOY,?\s*"
        r"to (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)",
        re.IGNORECASE,
    ),
    # "The average price of EU steel HRC increased by 6.4% MOM and rose by
    # 10.6% YOY to €705/MT" — no "as of {Month} {Day}" anchor, just "The
    # average {mod} price of X ..."
    re.compile(
        r"The average(?:\s+[a-zA-Z0-9 \-()]+?)? price of (?P<name>[A-Z][^.]+?) "
        r"(?P<dir1>increased|decreased|rose|declined|fell) by (?P<mom>\d+(?:\.\d+)?)%\s*MOM "
        r"and (?P<dir2>increased|decreased|rose|declined|fell) by (?P<yoy>\d+(?:\.\d+)?)%\s*YOY "
        r"to (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)",
        re.IGNORECASE,
    ),
    # Plural-prices "unchanged" form (apricots): "Turkish dried apricot
    # prices were unchanged MOM in March at $9,250/MT, up 85% YOY"
    re.compile(
        r"(?P<name>[A-Z][a-zA-Z]+(?: [a-zA-Z]+){1,5}) prices were unchanged MOM "
        r"(?:in (?P<as_of>[A-Za-z]+)\s*)?"
        r"at (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+),?\s*"
        r"(?P<dir2>up|down) (?P<yoy>\d+(?:\.\d+)?)%\s*YOY",
        re.IGNORECASE,
    ),
]


def _parse_match(m, pattern_idx: int) -> dict:
    d = m.groupdict()
    neg = {"decreased", "declined", "fell", "down"}
    mom_pct = 0.0 if "mom" not in d else (
        float(d["mom"]) * (-1 if d["dir1"].lower() in neg else 1)
    )
    if "mom" not in d:
        # "was unchanged MOM" patterns
        mom_pct = 0.0
    yoy_pct = float(d["yoy"]) * (-1 if d["dir2"].lower() in neg else 1)
    as_of = (d.get("as_of") or "").strip()
    return {
        "name": d["name"].strip(),
        "as_of": as_of,
        "mom_pct": mom_pct,
        "yoy_pct": yoy_pct,
        "price": f"{d['price']}/{d['unit']}".strip(),
    }


def extract_commentary(text: str, summary_rows: list[dict] | None = None) -> list[dict]:
    """Pull standardized "The average weekly price of X..." lines out of the
    body. Text should be column-split (see pdf_columns_to_text) so paragraphs
    aren't interleaved.

    If `summary_rows` is passed, tries to resolve each commentary's free-form
    name (e.g. "CBOT Maize" or "US Soyabean") to the canonical summary-table
    name (e.g. "Maize CBOT" or "Soyabean US") via token-set matching.
    """
    paragraphs: list[str] = []
    buf: list[str] = []
    for line in text.splitlines():
        if line.strip() == "":
            if buf:
                paragraphs.append(" ".join(s.strip() for s in buf))
                buf = []
        else:
            buf.append(line)
    if buf:
        paragraphs.append(" ".join(s.strip() for s in buf))

    def tokenise(s: str) -> list[str]:
        """Lowercase + drop trailing '*' (used to flag alt-period data) +
        strip trailing 's' so 'hazelnuts' and 'hazelnut' compare equal."""
        out = []
        for t in s.lower().split():
            t = t.rstrip("*").rstrip(",").rstrip(".")
            if not t:
                continue
            # Strip plural 's' but keep intentional abbreviations/codes
            # (ICE, EBP, LME, US, EU, UK, etc. stay 2-3 char uppercase).
            if len(t) > 3 and t.endswith("s") and not t.endswith("ss"):
                t = t[:-1]
            out.append(t)
        return out

    # Token-set index of summary names for fuzzy matching
    summary_index: dict[frozenset[str], str] = {}
    if summary_rows:
        for r in summary_rows:
            key = frozenset(tokenise(r["name"]))
            summary_index[key] = r["name"]

    # Commentary uses country adjectives ("Indian Rice"); summary uses the
    # country noun ("Rice India"). Normalise so token-set matching works.
    ADJ_TO_COUNTRY = {
        "indian": "india", "vietnamese": "vietnam", "chinese": "china",
        "russian": "russia", "ukrainian": "ukraine", "argentinian": "argentina",
        "brazilian": "brazil", "canadian": "canada", "spanish": "spain",
        "thai": "thailand", "australian": "australia", "norwegian": "norway",
        "chilean": "chile", "polish": "poland", "german": "germany",
        "french": "france", "italian": "italy", "dutch": "netherlands",
        "british": "uk", "turkish": "turkey", "malagasy": "madagascar",
        "malaysian": "malaysia", "indonesian": "indonesia",
    }

    def normalise_tokens(tokens: list[str]) -> list[str]:
        return [ADJ_TO_COUNTRY.get(t, t) for t in tokens]

    # Multi-word adjective → noun rewrites applied before tokenising
    MULTI_ADJ = [
        (r"\bsouth african\b", "south africa"),
        (r"\bnew zealander\b", "new zealand"),
    ]

    def canonicalise(free_name: str) -> str | None:
        if not summary_index:
            return None
        low = free_name.lower()
        for pat, repl in MULTI_ADJ:
            low = re.sub(pat, repl, low)
        tokens = normalise_tokens(tokenise(low))
        free_set = set(tokens)

        key = frozenset(tokens)
        if key in summary_index:
            return summary_index[key]
        # Subset match: summary tokens ⊆ commentary tokens
        # (handles "spot ICE London #5 White Sugar futures" → "Sugar ICE #5 London")
        best_match = None
        best_score = 0.0
        for row_tokens, row_name in summary_index.items():
            if row_tokens.issubset(free_set) and len(row_tokens) > best_score:
                best_match = row_name
                best_score = len(row_tokens)
        if best_match:
            return best_match
        # Jaccard-style overlap: when neither is a subset, pick the summary
        # row that shares the most tokens provided they share ≥ 2 meaningful
        # tokens (handles "ICE London Cocoa futures" → "Cocoa Bean ICE London").
        STOPS = {"the", "of", "and", "a", "an", "to", "in"}
        best_match = None
        best_score = 0.0
        for row_tokens, row_name in summary_index.items():
            shared = (row_tokens & free_set) - STOPS
            if len(shared) < 2:
                continue
            score = len(shared) / max(len(row_tokens - STOPS), 1)
            if score > best_score and score >= 0.5:
                best_match = row_name
                best_score = score
        if best_match:
            return best_match
        # Substring containment as last resort
        joined = " ".join(tokens)
        for row_name in summary_index.values():
            rn_low = row_name.lower()
            if joined in rn_low or rn_low in joined:
                return row_name
        return None

    neg = {"decreased", "declined", "fell"}
    out: list[dict] = []
    seen: set[str] = set()
    for p in paragraphs:
        # Try every pattern; record each match once (keyed by name) — a
        # commodity's narrative only belongs to one paragraph.
        for idx, pat in enumerate(PRICE_LINE_PATTERNS):
            for m in pat.finditer(p):
                parsed = _parse_match(m, idx)
                name = parsed["name"]
                if name in seen:
                    continue
                seen.add(name)
                out.append({
                    **parsed,
                    "canonical_name": canonicalise(name) or name,
                    "paragraph": p.strip(),
                })
    return out


# ---------- HICP ----------

def extract_hicp(hicp_xlsx: Path) -> dict:
    wb = openpyxl.load_workbook(hicp_xlsx, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    # Header row: every 2nd column from index 1 has a YYYY-MM string
    header = rows[0]
    months: list[str] = [header[i] for i in range(1, len(header), 2) if header[i]]

    # Rows 3..end are country data (row 2 is blank "GEO (Labels)")
    series: list[dict] = []
    for r in rows[2:]:
        label = r[0]
        if not label:
            continue
        values: list[float | None] = []
        for i in range(1, len(r), 2):
            v = r[i]
            if v is None or v == "" or v == ":":
                values.append(None)
            else:
                try:
                    values.append(float(v))
                except (TypeError, ValueError):
                    values.append(None)
        # Truncate to number of months
        values = values[: len(months)]
        # Shorten the EU label
        name = label
        if name.startswith("European Union"):
            name = "European Union"
        series.append({"geo": name, "values": values})

    return {
        "unit": "HICP YoY % change (Food, alcohol & tobacco)",
        "months": months,
        "series": series,
    }


# ---------- Forecast ----------

def _iter_data_sheet(ws) -> tuple[list[str], list[tuple]]:
    """Return (headers, data_rows) skipping disclaimer rows and blanks."""
    rows = list(ws.iter_rows(values_only=True))
    # Find the header row: first row whose first cell == "Date"
    header_idx = None
    for i, r in enumerate(rows):
        if r and r[0] == "Date":
            header_idx = i
            break
    if header_idx is None:
        return [], []
    header = [c for c in rows[header_idx] if c is not None]
    data = []
    for r in rows[header_idx + 1:]:
        if r[0] is None:
            continue
        data.append(r[: len(header)])
    return header, data


def extract_forecast(forecast_xlsx: Path) -> dict:
    wb = openpyxl.load_workbook(forecast_xlsx, data_only=True)
    commodities: list[dict] = []
    for sheet_name in wb.sheetnames:
        header, rows = _iter_data_sheet(wb[sheet_name])
        if not header:
            continue
        # header[0] is "Date"; remaining are commodity labels
        for col_idx, label in enumerate(header[1:], start=1):
            points = []
            for r in rows:
                d = r[0]
                v = r[col_idx] if col_idx < len(r) else None
                if not isinstance(d, (date, datetime)) or v is None:
                    continue
                points.append({"date": d.date().isoformat(), "value": float(v)})
            if not points:
                continue
            # Split Mintec code / description / unit from header label,
            # e.g. "COCL - Cocoa bean London ICE GBP/MT (L)"
            code = label.split(" - ", 1)[0].strip() if " - " in label else label
            description = label.split(" - ", 1)[1].strip() if " - " in label else label
            unit = ""
            m = re.search(r"([A-Z]{3}(?:/[A-Za-z/]+)?|[A-Z]+[/]?[A-Za-z]+)\s*\([LR]\)\s*$", label)
            if m:
                unit = m.group(1)
            commodities.append(
                {
                    "sheet": sheet_name,
                    "code": code,
                    "label": label,
                    "description": description,
                    "unit": unit,
                    "points": points,
                    "start": points[0]["date"],
                    "end": points[-1]["date"],
                }
            )
    return {
        "source": "Expana (Mintec) forward curves",
        "commodities": commodities,
    }


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit("usage: python scripts/extract.py <period-slug>  (e.g. 2026-04)")
    period = sys.argv[1]
    p = paths_for(period)
    if not p["raw"].exists():
        sys.exit(f"No raw/ folder at {p['raw']}")

    # Meta (for period descriptor). Optional — falls back to the period slug.
    meta_path = p["base"] / "meta.json"
    period_info = {"slug": period}
    if meta_path.exists():
        m = json.loads(meta_path.read_text())
        period_info = {
            "slug": period,
            "period": m.get("period"),
            "mom": m.get("period_mom"),
            "yoy": m.get("period_yoy"),
        }

    outputs: dict[str, int | str] = {}

    if p["pdf"].exists():
        # Summary table: full-page text works (each row fits on a single
        # line with both columns' % pairs visible).
        text = pdf_to_text(p["pdf"])
        summary = extract_summary(text)
        # Commentary: paragraphs span multiple lines and straddle the two
        # columns — so we crop each column separately to avoid interleaving.
        columns_text = pdf_columns_to_text(p["pdf"])
        commentary = extract_commentary(columns_text, summary)
        (p["base"] / "commodities.json").write_text(
            json.dumps({"period": period_info, "rows": summary}, indent=2)
        )
        (p["base"] / "commentary.json").write_text(
            json.dumps({"period": period_info, "entries": commentary}, indent=2)
        )
        outputs["commodities"] = len(summary)
        outputs["commentary"] = len(commentary)

    if p["hicp_xlsx"].exists():
        hicp = extract_hicp(p["hicp_xlsx"])
        (p["base"] / "hicp.json").write_text(json.dumps(hicp, indent=2))
        outputs["hicp"] = f"{len(hicp['series'])} series x {len(hicp['months'])} months"

    if p["forecast_xlsx"].exists():
        forecast = extract_forecast(p["forecast_xlsx"])
        (p["base"] / "forecast.json").write_text(json.dumps(forecast, indent=2))
        outputs["forecast"] = (
            f"{len(forecast['commodities'])} commodities; "
            f"{sum(len(c['points']) for c in forecast['commodities'])} points"
        )

    for k, v in outputs.items():
        print(f"{k:<12} {v}")


if __name__ == "__main__":
    main()
