"""Extract April 2026 report data into JSON.

Reads the three source files at the repo root:
  - Commodity Price Change Overview Report April 2026.pdf
  - HICP Data EU.xlsx
  - Forecast Data Q1 26.xlsx

Writes JSON under data/april-2026/:
  - commodities.json    Summary table (category, name, MoM, YoY) for all ~117 commodities
  - commentary.json     Per-commodity price line + narrative paragraphs extracted from PDF
  - hicp.json           EU and country-level HICP food inflation YoY by month
  - forecast.json       Daily forward curves for 12 commodities (May 2026+)

Re-run: python scripts/extract_april_2026.py
"""
from __future__ import annotations

import json
import re
import subprocess
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PERIOD_DIR = ROOT / "data" / "2026-04"
RAW = PERIOD_DIR / "raw"
PDF = RAW / "overview.pdf"
HICP_XLSX = RAW / "hicp.xlsx"
FORECAST_XLSX = RAW / "forecast.xlsx"
OUT = PERIOD_DIR
OUT.mkdir(parents=True, exist_ok=True)

PERIOD = {
    "reportMonth": "April 2026",
    "mom": "March 2026 vs February 2026",
    "yoy": "March 2026 vs March 2025",
}

CATEGORIES = [
    "Grains & Feed",
    "Oilseeds & Vegetable Oils",
    "Meat & Poultry",
    "Dairy & Eggs",
    "Fish & Seafood",
    "Fruit & Vegetables",
    "Juices",
    "Softs",
    "Herbs & Spices",
    "Nuts & Dried Fruit",
    "Textiles",
    "Packaging",
]


def pdf_to_text() -> str:
    out = subprocess.check_output(
        ["pdftotext", "-layout", str(PDF), "-"], text=True, encoding="utf-8"
    )
    return out


# ---------- Summary table extraction ----------

PCT = re.compile(r"([+\-]?\d+(?:\.\d+)?)%")
ROW_2PCT = re.compile(r"([+\-]?\d+(?:\.\d+)?%)\s+([+\-]?\d+(?:\.\d+)?%)")


def parse_pct(s: str) -> float | None:
    m = PCT.search(s)
    return float(m.group(1)) if m else None


def extract_summary(text: str) -> list[dict]:
    """Parse the 4-page summary table into {category, name, mom, yoy} rows.

    Layout: two side-by-side columns. Each row ends with `...MoM%  YoY%`.
    Category labels appear centered in a separate mini-column; we assign by
    running through categories in the order they appear in the document.
    """
    lines = text.splitlines()
    rows: list[dict] = []

    # Category runs: iterate through lines; when a category appears at the start
    # of a column, remember it for that column.
    # Simpler approach: scan for rows with two percentages. Split line at col
    # midpoint to pick name. Track category via preceding category-labeled rows.
    left_cat: str | None = None
    right_cat: str | None = None

    # We'll find the column midpoint by locating the gap between columns on
    # the header line(s). Hardcode to ~80 based on layout.
    MID = 80

    for line in lines:
        stripped = line.rstrip()
        if not stripped:
            continue
        # Category detection: look for exact category tokens anywhere on line
        for cat in CATEGORIES:
            if cat in line:
                col = line.index(cat)
                if col < MID:
                    left_cat = cat
                else:
                    right_cat = cat
        # Find rows that end with two percentages
        for m in ROW_2PCT.finditer(line):
            end = m.end()
            # The text preceding the match is the commodity name
            start_of_match = m.start()
            prefix = line[:start_of_match].rstrip()
            # Column side based on match position
            col = start_of_match
            cat = left_cat if col < MID else right_cat
            # Clean the commodity name: strip any category label, numeric noise
            name = prefix
            if cat and cat in name:
                name = name.replace(cat, "")
            # Collapse runs of whitespace; take trailing words (name is rightmost)
            name = " ".join(name.split())
            # Heuristic: commodity name is the trailing token group after the
            # category column. Drop a leading category if still present.
            for c in CATEGORIES:
                if name.startswith(c):
                    name = name[len(c):].strip()
            # Skip table header row
            if not name or name.lower().startswith("commodity"):
                continue
            mom_s, yoy_s = m.group(1), m.group(2)
            rows.append(
                {
                    "category": cat or "Unknown",
                    "name": name,
                    "mom_pct": parse_pct(mom_s),
                    "yoy_pct": parse_pct(yoy_s),
                }
            )
    # Dedupe (same row might appear if layout detection double-matches)
    seen = set()
    out: list[dict] = []
    for r in rows:
        key = (r["category"], r["name"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


# ---------- Per-commodity commentary extraction ----------

PRICE_LINE = re.compile(
    r"The average weekly price of (?P<name>[A-Z][^.]+?) as of (?P<as_of>[A-Za-z]+ \d+) "
    r"(?P<dir1>increased|decreased|rose|declined|fell) by (?P<mom>\d+(?:\.\d+)?)% MOM "
    r"and (?P<dir2>increased|decreased|rose|declined|fell) by (?P<yoy>\d+(?:\.\d+)?)% YOY "
    r"to (?P<price>[^/\s]+(?:[ ,][^/\s]+)?)/(?P<unit>[A-Za-z0-9]+)",
    re.IGNORECASE,
)


def extract_commentary(text: str) -> list[dict]:
    """Pull standardized "The average weekly price of X..." lines out of the
    body. The PDF renders paragraphs in two side-by-side columns, so a single
    paragraph block from pdftotext may contain two commodity narratives.
    """
    paragraphs: list[str] = []
    buf: list[str] = []
    for line in text.splitlines():
        if line.strip() == "":
            if buf:
                paragraphs.append(" ".join(s.strip() for s in buf))
                buf = []
        else:
            buf.append(line)
    if buf:
        paragraphs.append(" ".join(s.strip() for s in buf))

    neg = {"decreased", "declined", "fell"}
    out: list[dict] = []
    seen: set[str] = set()
    for p in paragraphs:
        for m in PRICE_LINE.finditer(p):
            d = m.groupdict()
            name = d["name"].strip()
            if name in seen:
                continue
            seen.add(name)
            out.append(
                {
                    "name": name,
                    "as_of": d["as_of"].strip(),
                    "mom_pct": float(d["mom"]) * (-1 if d["dir1"].lower() in neg else 1),
                    "yoy_pct": float(d["yoy"]) * (-1 if d["dir2"].lower() in neg else 1),
                    "price": f"{d['price']}/{d['unit']}".strip(),
                    "paragraph": p.strip(),
                }
            )
    return out


# ---------- HICP ----------

def extract_hicp() -> dict:
    wb = openpyxl.load_workbook(HICP_XLSX, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    # Header row: every 2nd column from index 1 has a YYYY-MM string
    header = rows[0]
    months: list[str] = [header[i] for i in range(1, len(header), 2) if header[i]]

    # Rows 3..end are country data (row 2 is blank "GEO (Labels)")
    series: list[dict] = []
    for r in rows[2:]:
        label = r[0]
        if not label:
            continue
        values: list[float | None] = []
        for i in range(1, len(r), 2):
            v = r[i]
            if v is None or v == "" or v == ":":
                values.append(None)
            else:
                try:
                    values.append(float(v))
                except (TypeError, ValueError):
                    values.append(None)
        # Truncate to number of months
        values = values[: len(months)]
        # Shorten the EU label
        name = label
        if name.startswith("European Union"):
            name = "European Union"
        series.append({"geo": name, "values": values})

    return {
        "unit": "HICP YoY % change (Food, alcohol & tobacco)",
        "months": months,
        "series": series,
    }


# ---------- Forecast ----------

def _iter_data_sheet(ws) -> tuple[list[str], list[tuple]]:
    """Return (headers, data_rows) skipping disclaimer rows and blanks."""
    rows = list(ws.iter_rows(values_only=True))
    # Find the header row: first row whose first cell == "Date"
    header_idx = None
    for i, r in enumerate(rows):
        if r and r[0] == "Date":
            header_idx = i
            break
    if header_idx is None:
        return [], []
    header = [c for c in rows[header_idx] if c is not None]
    data = []
    for r in rows[header_idx + 1:]:
        if r[0] is None:
            continue
        data.append(r[: len(header)])
    return header, data


def extract_forecast() -> dict:
    wb = openpyxl.load_workbook(FORECAST_XLSX, data_only=True)
    commodities: list[dict] = []
    for sheet_name in wb.sheetnames:
        header, rows = _iter_data_sheet(wb[sheet_name])
        if not header:
            continue
        # header[0] is "Date"; remaining are commodity labels
        for col_idx, label in enumerate(header[1:], start=1):
            points = []
            for r in rows:
                d = r[0]
                v = r[col_idx] if col_idx < len(r) else None
                if not isinstance(d, (date, datetime)) or v is None:
                    continue
                points.append({"date": d.date().isoformat(), "value": float(v)})
            if not points:
                continue
            # Split Mintec code / description / unit from header label,
            # e.g. "COCL - Cocoa bean London ICE GBP/MT (L)"
            code = label.split(" - ", 1)[0].strip() if " - " in label else label
            description = label.split(" - ", 1)[1].strip() if " - " in label else label
            unit = ""
            m = re.search(r"([A-Z]{3}(?:/[A-Za-z/]+)?|[A-Z]+[/]?[A-Za-z]+)\s*\([LR]\)\s*$", label)
            if m:
                unit = m.group(1)
            commodities.append(
                {
                    "sheet": sheet_name,
                    "code": code,
                    "label": label,
                    "description": description,
                    "unit": unit,
                    "points": points,
                    "start": points[0]["date"],
                    "end": points[-1]["date"],
                }
            )
    return {
        "source": "Expana (Mintec) forward curves - Q1 2026 export",
        "commodities": commodities,
    }


def main() -> None:
    text = pdf_to_text()

    summary = extract_summary(text)
    commentary = extract_commentary(text)
    hicp = extract_hicp()
    forecast = extract_forecast()

    (OUT / "commodities.json").write_text(
        json.dumps({"period": PERIOD, "rows": summary}, indent=2)
    )
    (OUT / "commentary.json").write_text(
        json.dumps({"period": PERIOD, "entries": commentary}, indent=2)
    )
    (OUT / "hicp.json").write_text(json.dumps(hicp, indent=2))
    (OUT / "forecast.json").write_text(json.dumps(forecast, indent=2))

    print(f"commodities.json : {len(summary)} rows")
    print(f"commentary.json  : {len(commentary)} entries")
    print(f"hicp.json        : {len(hicp['series'])} series over {len(hicp['months'])} months")
    print(
        f"forecast.json    : {len(forecast['commodities'])} commodities; "
        f"total points = {sum(len(c['points']) for c in forecast['commodities'])}"
    )


if __name__ == "__main__":
    main()
